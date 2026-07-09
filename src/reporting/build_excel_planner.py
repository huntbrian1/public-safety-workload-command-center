from __future__ import annotations

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src import config
from src.utils.logging_utils import get_logger

logger = get_logger(__name__)


HEADER_FILL = PatternFill("solid", fgColor="1F4E5F")
SECTION_FILL = PatternFill("solid", fgColor="E7F0F3")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=15, bold=True, color="1F4E5F")
THIN = Side(style="thin", color="D9E2E7")


def _read_csv(name: str) -> pd.DataFrame:
    path = config.HEX_OUTPUT_DIR / name
    return pd.read_csv(path) if path.exists() else pd.DataFrame()


def _write_df(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1, max_rows: int | None = None):
    if max_rows is not None:
        df = df.head(max_rows)
    if df.empty:
        ws.cell(start_row, start_col, "No data generated in this run.")
        return
    for j, col in enumerate(df.columns, start_col):
        cell = ws.cell(start_row, j, col)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center")
    for i, row in enumerate(df.itertuples(index=False), start_row + 1):
        for j, value in enumerate(row, start_col):
            ws.cell(i, j, value)
    end_row = start_row + len(df)
    end_col = start_col + len(df.columns) - 1
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = Border(bottom=THIN)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.auto_filter.ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"


def _format_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    for col in range(1, min(ws.max_column, 12) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 26


def _set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def build_excel_planner() -> None:
    config.ensure_directories()
    kpis = _read_csv("hex_executive_kpis.csv")
    zones = _read_csv("hex_zone_summary.csv")
    preds = _read_csv("hex_model_predictions.csv")
    recs = _read_csv("hex_recommendations.csv")
    quality = _read_csv("hex_data_quality_summary.csv")

    wb = Workbook()
    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "Public Safety Service Demand Intelligence Scenario Planner"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "Use this workbook to review executive KPIs, zone workload-risk scores, forecast outputs, and resource-planning scenarios. Yellow cells on the Capacity Assumptions sheet are editable."
    ws["A3"].alignment = Alignment(wrap_text=True)
    ws["A5"] = "Planning framing"
    ws["A5"].fill = SECTION_FILL
    ws["A6"] = "The workbook supports aggregate operational workload planning and product/service analytics. It should not be used for individual-level decisions."
    ws["A6"].alignment = Alignment(wrap_text=True)
    ws.column_dimensions["A"].width = 105
    ws.row_dimensions[3].height = 42
    ws.row_dimensions[6].height = 42

    ws = wb.create_sheet("Executive KPIs")
    _write_df(ws, kpis, max_rows=50)
    _format_sheet(ws)
    _set_widths(ws, {"A": 28, "B": 24, "C": 54})
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 34

    ws = wb.create_sheet("Zone Summary")
    _write_df(ws, zones, max_rows=100)
    if not zones.empty and "total_events" in zones.columns:
        chart = BarChart()
        chart.title = "Top Zones by Service Demand"
        chart.y_axis.title = "Events"
        chart.x_axis.title = "Zone"
        data = Reference(ws, min_col=list(zones.columns).index("total_events") + 1, min_row=1, max_row=min(11, len(zones) + 1))
        cats = Reference(ws, min_col=1, min_row=2, max_row=min(11, len(zones) + 1))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 7
        chart.width = 14
        ws.add_chart(chart, "J2")
    _format_sheet(ws)
    _set_widths(ws, {"A": 12, "B": 14, "C": 20, "D": 20, "E": 18, "F": 18, "G": 18, "H": 16, "I": 16, "J": 16, "K": 16, "L": 16, "M": 16})

    ws = wb.create_sheet("Forecast Output")
    _write_df(ws, preds, max_rows=1000)
    _format_sheet(ws)
    _set_widths(ws, {"A": 12, "B": 22, "C": 18, "D": 18, "E": 18, "F": 16, "G": 22, "H": 18})

    ws = wb.create_sheet("Capacity Assumptions")
    assumptions = [
        ("available_staff_resources", 30, "Total resources available for the scenario"),
        ("target_demand_coverage_pct", 0.85, "Target share of projected workload to cover"),
        ("expected_demand_increase_pct", 0.10, "Planning increase over observed demand"),
        ("weather_event_demand_multiplier", 1.05, "Multiplier for weather/event-sensitive periods"),
        ("high_risk_threshold", 70, "Risk score threshold for high-risk zones"),
        ("max_capacity_per_resource", 12, "Demand units one resource can cover"),
        ("minimum_coverage_target", 0.75, "Minimum acceptable projected coverage"),
    ]
    ws.append(["Input", "Value", "Description"])
    for row in assumptions:
        ws.append(row)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
    for row in range(2, len(assumptions) + 2):
        ws.cell(row, 2).fill = INPUT_FILL
    ws["B3"].number_format = "0%"
    ws["B4"].number_format = "0%"
    ws["B5"].number_format = "0.00"
    ws["B8"].number_format = "0%"
    _format_sheet(ws)
    _set_widths(ws, {"A": 34, "B": 18, "C": 62})
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 32

    ws = wb.create_sheet("Scenario Planner")
    headers = [
        "zone_id",
        "baseline_events",
        "workload_risk_score",
        "priority_weight",
        "projected_demand",
        "resources_needed",
        "allocated_resources",
        "projected_capacity",
        "workload_gap",
        "coverage_pct",
        "recommendation_text",
    ]
    ws.append(headers)
    top_zones = zones.head(25).copy() if not zones.empty else pd.DataFrame(columns=["zone_id", "total_events", "workload_risk_score"])
    for idx, row in enumerate(top_zones.itertuples(index=False), start=2):
        zone = getattr(row, "zone_id", "")
        total_events = float(getattr(row, "total_events", 0) or 0)
        risk = float(getattr(row, "workload_risk_score", 0) or 0)
        ws.cell(idx, 1, zone)
        ws.cell(idx, 2, total_events)
        ws.cell(idx, 3, risk)
        ws.cell(idx, 4, f"=C{idx}/SUM($C$2:$C${len(top_zones)+1})")
        ws.cell(idx, 5, f"=B{idx}*(1+'Capacity Assumptions'!$B$4)*'Capacity Assumptions'!$B$5")
        ws.cell(idx, 6, f"=E{idx}*'Capacity Assumptions'!$B$3/'Capacity Assumptions'!$B$7")
        ws.cell(idx, 7, f"='Capacity Assumptions'!$B$2*D{idx}")
        ws.cell(idx, 8, f"=G{idx}*'Capacity Assumptions'!$B$7")
        ws.cell(idx, 9, f"=MAX(0,E{idx}*'Capacity Assumptions'!$B$3-H{idx})")
        ws.cell(idx, 10, f"=IF(E{idx}=0,0,H{idx}/E{idx})")
        ws.cell(idx, 11, f'=IF(C{idx}>=\'Capacity Assumptions\'!$B$6,"Prioritize workload-risk review","Monitor recurring demand pattern")')
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
    for col in ["D", "J"]:
        for row in range(2, len(top_zones) + 2):
            ws[f"{col}{row}"].number_format = "0.0%"
    _format_sheet(ws)
    _set_widths(ws, {"A": 12, "B": 16, "C": 18, "D": 16, "E": 18, "F": 18, "G": 18, "H": 18, "I": 16, "J": 14, "K": 38})
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 30

    ws = wb.create_sheet("Recommendations")
    _write_df(ws, recs, max_rows=100)
    _format_sheet(ws)
    _set_widths(ws, {"A": 14, "B": 30, "C": 30, "D": 46, "E": 52, "F": 12})
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 58

    ws = wb.create_sheet("Data Quality Notes")
    _write_df(ws, quality, max_rows=200)
    _format_sheet(ws)
    _set_widths(ws, {"A": 38, "B": 18, "C": 58})
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 34

    out_path = config.EXCEL_OUTPUT_DIR / "public_safety_scenario_planner.xlsx"
    wb.save(out_path)
    check = load_workbook(out_path, data_only=False)
    missing = [name for name in ["Instructions", "Executive KPIs", "Zone Summary", "Forecast Output", "Capacity Assumptions", "Scenario Planner", "Recommendations", "Data Quality Notes"] if name not in check.sheetnames]
    if missing:
        raise RuntimeError(f"Workbook missing expected sheets: {missing}")
    logger.info("Excel scenario planner written to %s", out_path)


if __name__ == "__main__":
    build_excel_planner()
